# -*- coding: utf-8 -*-
import openpyxl
from sortList import sort_list


def pro_xls(sourcePath):
    wb=openpyxl.load_workbook(sourcePath)
    names=[]
    bm=0#部门所在列，索引从0开始
    jb=0#级别所在列，索引从0开始
    for sheet in wb.worksheets:#获取不重复的姓名列表
        for r in range(1,sheet.max_row):
            name=sheet.cell(row=r+1,column=1).value
            if name==None:
                continue
            else:
                names.append(name.strip())
        names=list(set(names))
    wb.create_sheet('汇总')#建立汇总表
    for c in range(wb.worksheets[0].max_column):#设置汇总表头
        v=wb.worksheets[0].cell(row=1,column=c+1).value.strip()
        wb['汇总'].cell(row=1,column=c+1).value=v
        if v=='部门':
            bm=c
        if v=='级别':
            jb=c

    temp_list = []
    for n in names:#按姓名遍历所有表格，并按姓名汇总所有数字单元格
        temp_values = []
        for sheet in wb.worksheets:

            for r in range(1, sheet.max_row):
                name=sheet.cell(row=r + 1, column=1).value
                if name==None:
                    continue
                else:
                    name=name.strip()
                if n==name:
                    dist_row=sheet.cell(row=r + 1, column=1).row
                    for c in range(sheet.max_column):
                        tv=sheet.cell(row=dist_row, column=c + 1).value
                        if len(temp_values)<sheet.max_column:
                            temp_values.append(tv)
                        else:
                            if type(tv)==int or type(tv)==float:
                                if temp_values[c]==None:
                                    temp_values[c]=tv
                                else:
                                    temp_values[c]+=tv
                            else:
                                if tv!=None:#如果不是数字，并且不为空则直接替换之前的值
                                    temp_values[c] = tv.strip()
                    break
        if temp_values[jb]=='中层正职':#中层正职改为中层
            temp_values[jb]='中层'
        temp_list.append(temp_values)
    sorted_list=sort_list(temp_list,bm,jb)
    for row in sorted_list:
        wb['汇总'].append(row)
    # wb.save(distPath)
    return wb



